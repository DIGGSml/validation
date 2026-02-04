#!/usr/bin/env python3
"""
DIGGS Schema Compatibility Analyzer
Analyzes backward compatibility between two versions of DIGGS schemas

Author: Dan Ponti / Claude (Anthropic)
Version: 1.0.0
"""

import os
import sys
import argparse
import re
from pathlib import Path
from typing import Dict, List, Tuple, Set, Optional
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Import the schema resolver utilities
from fixed_resolver import (
    parse_schema, get_all_types, get_all_attributes, get_all_simpletypes,
    resolve_content_model, get_base_type_name,
    NS_26, NS_3
)


class DIGGSCompatibilityAnalyzer:
    """Main analyzer class for DIGGS schema compatibility checking"""
    
    def __init__(self, old_dir: str, new_dir: str, mapping_file: Optional[str] = None):
        self.old_dir = Path(old_dir)
        self.new_dir = Path(new_dir)
        self.mapping_file = Path(mapping_file) if mapping_file else None
        
        self.old_version = None
        self.new_version = None
        self.type_mappings = {}
        self.type_compat_cache = {}
        
        # Schema collections
        self.old_schemas = []
        self.new_schemas = []
        self.old_types = {}
        self.old_attrs = {}
        self.old_simpletypes = {}
        self.new_types = {}
        self.new_attrs = {}
        self.new_simpletypes = {}
        
    def detect_version(self, directory: Path) -> str:
        """Detect DIGGS version from schema files in directory"""
        versions = set()
        
        for xsd_file in directory.rglob('*.xsd'):
            try:
                with open(xsd_file, 'r', encoding='utf-8') as f:
                    content = f.read(5000)  # Read first 5KB to ensure we get schema element
                    
                    # Find schema element(s)
                    schema_pattern = re.compile(r'<schema[^>]*>', re.IGNORECASE | re.DOTALL)
                    matches = schema_pattern.finditer(content)
                    
                    for match in matches:
                        schema_tag = match.group(0)
                        
                        # Only process schemas with diggsml.org in targetNamespace
                        if 'diggsml.org' in schema_tag:
                            # Extract version from this schema element
                            version_match = re.search(r'version\s*=\s*["\']([^"\']+)["\']', schema_tag)
                            if version_match:
                                version = version_match.group(1)
                                # Skip XML declaration versions (1.0, 1.1)
                                if version not in ['1.0', '1.1']:
                                    versions.add(version)
                                    break  # Found DIGGS version in this file
                                    
            except Exception as e:
                print(f"Warning: Could not read {xsd_file}: {e}")
        
        if len(versions) == 0:
            raise ValueError(f"No DIGGS schema version found in {directory}. "
                           f"Make sure directory contains .xsd files with targetNamespace='...diggsml.org...'")
        elif len(versions) > 1:
            print(f"Warning: Multiple DIGGS versions found in {directory}: {versions}")
            print(f"Using version: {sorted(versions)[0]}")
        
        return sorted(versions)[0]
    
    def load_mappings(self) -> Dict[str, str]:
        """Load type mappings from tab-separated file"""
        mappings = {}
        
        if not self.mapping_file or not self.mapping_file.exists():
            return mappings
        
        print(f"\nLoading type mappings from: {self.mapping_file}")
        
        try:
            with open(self.mapping_file, 'r', encoding='utf-8') as f:
                for line_num, line in enumerate(f, 1):
                    line = line.strip()
                    if not line or line.startswith('#'):
                        continue
                    
                    parts = line.split('\t')
                    if len(parts) >= 2:
                        old_type = parts[0].strip()
                        new_type = parts[1].strip()
                        if old_type and new_type:
                            mappings[old_type] = new_type
                    else:
                        print(f"Warning: Invalid mapping on line {line_num}: {line}")
            
            print(f"  Loaded {len(mappings)} type mappings")
        
        except Exception as e:
            print(f"Error loading mappings: {e}")
            return {}
        
        return mappings
    
    def find_schema_files(self, directory: Path) -> List[Path]:
        """Recursively find all .xsd files in directory"""
        return sorted(directory.rglob('*.xsd'))
    
    def load_schemas(self, directory: Path, namespace_dict: dict) -> Tuple[List, Dict, Dict, Dict]:
        """Load all schemas from directory"""
        print(f"\nLoading schemas from: {directory}")
        
        schema_files = self.find_schema_files(directory)
        if not schema_files:
            raise ValueError(f"No .xsd files found in {directory}")
        
        # Track schemas with their namespace labels
        schemas_with_ns = []
        
        for xsd_file in schema_files:
            # Extract namespace from schema's targetNamespace attribute
            relative_path = xsd_file.relative_to(directory)
            ns_label = self._extract_namespace_label(xsd_file)
            
            schema = parse_schema(str(xsd_file), namespace_dict)
            if schema is not None:
                schemas_with_ns.append((schema, ns_label))
                print(f"  ✓ {relative_path} ({ns_label})")
        
        # Collect types with qualified names
        all_types = {}
        all_simpletypes = {}
        
        for schema, ns_label in schemas_with_ns:
            # Collect complexTypes with namespace prefix
            for ct in schema.findall('.//xs:complexType', namespace_dict):
                name = ct.get('name')
                if name:
                    # Use qualified name as key
                    qualified_name = f"{ns_label}:{name}"
                    all_types[qualified_name] = ct
            
            # Collect simpleTypes with namespace prefix
            for st in schema.findall('.//xs:simpleType', namespace_dict):
                name = st.get('name')
                if name:
                    qualified_name = f"{ns_label}:{name}"
                    all_simpletypes[qualified_name] = st
        
        # Attributes don't need namespace qualification (they're referenced differently)
        schemas_only = [s for s, ns in schemas_with_ns]
        all_attrs = get_all_attributes(schemas_only, namespace_dict)
        
        print(f"  Total: {len(all_types)} complexTypes, {len(all_simpletypes)} simpleTypes, {len(all_attrs)} attributes")
        
        return schemas_with_ns, all_types, all_attrs, all_simpletypes
    
    def _extract_namespace_label(self, xsd_file: Path) -> str:
        """Extract namespace label from schema's targetNamespace attribute"""
        try:
            with open(xsd_file, 'r', encoding='utf-8') as f:
                content = f.read(3000)  # Read first 3KB to get schema element
                
                # Find targetNamespace attribute
                match = re.search(r'targetNamespace\s*=\s*["\']([^"\']+)["\']', content)
                if not match:
                    return 'unknown'
                
                target_ns = match.group(1)
                
                # Map targetNamespace to namespace label
                if 'diggsml.org/schemas' in target_ns:
                    if 'geotechnical' in target_ns:
                        return 'diggs_geo'
                    else:
                        return 'diggs'
                elif 'energistics.org/energyml/data/commonv2' in target_ns:
                    return 'eml'
                elif 'energistics.org/energyml/data/witsmlv2' in target_ns:
                    return 'witsml'
                elif 'opengis.net/gml' in target_ns:
                    if '/lrov' in target_ns:
                        return 'glrov'
                    elif '/lr' in target_ns:
                        return 'glr'
                    else:
                        return 'gml'
                else:
                    # For unknown namespaces, try to extract a reasonable label
                    # from the last part of the namespace
                    parts = target_ns.rstrip('/').split('/')
                    return parts[-1] if parts else 'unknown'
                    
        except Exception as e:
            print(f"Warning: Could not extract namespace from {xsd_file}: {e}")
            return 'unknown'
    
    def normalize_type_name(self, type_name: str) -> str:
        """Normalize type names - preserves namespace prefixes
        
        Uses clean_type_name from fixed_resolver to ensure consistent format
        """
        if not type_name:
            return ''
        from fixed_resolver import clean_type_name
        return clean_type_name(type_name)
    
    def is_builtin_type(self, type_name: str) -> bool:
        """Check if type is an XML Schema built-in type"""
        if not type_name:
            return False
        
        # Check if it's an xs: prefixed type
        if type_name.startswith('xs:'):
            return True
        
        # Check bare name for common built-ins
        if ':' in type_name:
            bare_name = type_name.split(':', 1)[1]
        else:
            bare_name = type_name
            
        builtins = ['string', 'double', 'float', 'integer', 'int', 'long', 'short', 
                   'byte', 'boolean', 'decimal', 'date', 'dateTime', 'time', 
                   'anyURI', 'QName', 'anyType', 'anySimpleType']
        return bare_name in builtins
    
    def check_type_content_compatibility(self, old_type: str, new_type: str, 
                                         depth: int = 0) -> Tuple[bool, str]:
        """
        Recursively check if changing from old_type to new_type is backward compatible
        by comparing their content models.
        
        Returns: (is_compatible, reason)
        """
        # Prevent infinite recursion
        if depth > 5:
            return (True, "Max recursion depth reached")
        
        # Check cache
        cache_key = (old_type, new_type)
        if cache_key in self.type_compat_cache:
            return self.type_compat_cache[cache_key]
        
        old_normalized = self.normalize_type_name(old_type)
        new_normalized = self.normalize_type_name(new_type)
        
        # If types are identical, compatible
        if old_normalized == new_normalized:
            result = (True, "Types identical")
            self.type_compat_cache[cache_key] = result
            return result
        
        # Check if there's a known mapping
        if old_normalized in self.type_mappings:
            mapped = self.normalize_type_name(self.type_mappings[old_normalized])
            if mapped == new_normalized:
                result = (True, "Known mapping")
                self.type_compat_cache[cache_key] = result
                return result
        
        # If either is a built-in type, we can't analyze further
        if self.is_builtin_type(old_type) or self.is_builtin_type(new_type):
            if old_normalized == new_normalized:
                result = (True, "Same built-in type")
            else:
                result = (False, f"Built-in type change: {old_normalized} → {new_normalized}")
            self.type_compat_cache[cache_key] = result
            return result
        
        # Both should be complex types - compare their content models
        if old_normalized not in self.old_types:
            result = (False, f"Old type {old_normalized} not found in old schema")
            self.type_compat_cache[cache_key] = result
            return result
        
        if new_normalized not in self.new_types:
            result = (False, f"New type {new_normalized} not found in new schema")
            self.type_compat_cache[cache_key] = result
            return result
        
        # Resolve content models for both types
        try:
            cm_old_elements, cm_old_attrs = resolve_content_model(
                old_normalized, self.old_types, self.old_attrs, NS_26
            )
            cm_new_elements, cm_new_attrs = resolve_content_model(
                new_normalized, self.new_types, self.new_attrs, NS_3
            )
        except Exception as e:
            result = (False, f"Error resolving content models: {str(e)}")
            self.type_compat_cache[cache_key] = result
            return result
        
        # Check for backward compatibility
        names_old = set(cm_old_elements.keys())
        names_new = set(cm_new_elements.keys())
        attr_names_old = set(cm_old_attrs.keys())
        attr_names_new = set(cm_new_attrs.keys())
        
        # Missing elements (incompatible)
        missing_elems = names_old - names_new
        if missing_elems:
            result = (False, f"Missing elements: {', '.join(sorted(list(missing_elems)[:3]))}")
            self.type_compat_cache[cache_key] = result
            return result
        
        # Missing attributes (incompatible)
        missing_attrs = attr_names_old - attr_names_new
        if missing_attrs:
            result = (False, f"Missing attributes: {', '.join(sorted(list(missing_attrs)[:3]))}")
            self.type_compat_cache[cache_key] = result
            return result
        
        # New required elements (incompatible)
        new_elems = names_new - names_old
        for elem_name in new_elems:
            if cm_new_elements[elem_name]['minOccurs'] != '0':
                result = (False, f"New required element: {elem_name}")
                self.type_compat_cache[cache_key] = result
                return result
        
        # New required attributes (incompatible)
        new_attrs = attr_names_new - attr_names_old
        for attr_name in new_attrs:
            if cm_new_attrs[attr_name]['use'] == 'required':
                result = (False, f"New required attribute: {attr_name}")
                self.type_compat_cache[cache_key] = result
                return result
        
        # Check cardinality restrictions (incompatible)
        common_elems = names_old & names_new
        for elem_name in common_elems:
            e_old = cm_old_elements[elem_name]
            e_new = cm_new_elements[elem_name]
            
            min_old = int(e_old['minOccurs']) if e_old['minOccurs'].isdigit() else 0
            min_new = int(e_new['minOccurs']) if e_new['minOccurs'].isdigit() else 0
            max_old = 999999 if e_old['maxOccurs'] == 'unbounded' else int(e_old['maxOccurs']) if e_old['maxOccurs'].isdigit() else 1
            max_new = 999999 if e_new['maxOccurs'] == 'unbounded' else int(e_new['maxOccurs']) if e_new['maxOccurs'].isdigit() else 1
            
            if min_new > min_old or max_new < max_old:
                result = (False, f"Cardinality restricted on element: {elem_name}")
                self.type_compat_cache[cache_key] = result
                return result
            
            # Recursively check if element type changes are compatible
            if e_old['type'] != e_new['type'] and e_new['type']:
                elem_type_compat, elem_reason = self.check_type_content_compatibility(
                    e_old['type'], e_new['type'], depth + 1
                )
                if not elem_type_compat:
                    result = (False, f"Incompatible type change on {elem_name}: {elem_reason}")
                    self.type_compat_cache[cache_key] = result
                    return result
        
        # Check attribute use restrictions (incompatible)
        common_attrs = attr_names_old & attr_names_new
        for attr_name in common_attrs:
            a_old = cm_old_attrs[attr_name]
            a_new = cm_new_attrs[attr_name]
            
            if a_old['use'] == 'optional' and a_new['use'] == 'required':
                result = (False, f"Attribute now required: {attr_name}")
                self.type_compat_cache[cache_key] = result
                return result
        
        # All checks passed - types are compatible!
        result = (True, "Content models compatible")
        self.type_compat_cache[cache_key] = result
        return result
    
    def is_type_change_compatible(self, old_type: str, new_type: str) -> Tuple[bool, str]:
        """Check if a type change is backward compatible"""
        if not old_type or not new_type:
            return (True, "Empty type")
        
        # First check simple cases
        old_normalized = self.normalize_type_name(old_type)
        new_normalized = self.normalize_type_name(new_type)
        
        if old_normalized == new_normalized:
            return (True, "Types identical")
        
        if old_normalized in self.type_mappings:
            mapped_type = self.normalize_type_name(self.type_mappings[old_normalized])
            if mapped_type == new_normalized:
                return (True, "Known mapping")
        
        # Now do deep content model comparison
        return self.check_type_content_compatibility(old_type, new_type, depth=0)
    
    def format_cardinality(self, min_occ: str, max_occ: str) -> str:
        """Format cardinality as (min..max)"""
        return f"({min_occ}..{max_occ})"
    
    def compare_type(self, qualified_name: str, mapped_name: Optional[str] = None) -> Dict:
        """Compare a single type between old and new versions
        
        Args:
            qualified_name: Qualified name like "diggs:AbstractFeature" or "eml:LengthMeasure"
            mapped_name: Optional mapped/renamed qualified name in new version
        """
        result = {
            'name': qualified_name,  # Now includes namespace prefix
            'type_removed': '',
            'base_changed': '',
            'structure_changes': '',
            'new_elements': [],
            'missing_elements': [],
            'expanded_cardinality': [],
            'restricted_cardinality': [],
            'type_changes': [],
            'notes': '',
            'compatible': 'Yes'
        }
        
        ct_old = self.old_types.get(qualified_name)
        ct_new = self.new_types.get(mapped_name if mapped_name else qualified_name)
        
        if ct_new is None:
            result['type_removed'] = 'Yes'
            result['compatible'] = 'No'
            result['notes'] = 'Type not found in new version'
            return result
        
        result['type_removed'] = 'No'
        
        if mapped_name:
            result['type_removed'] = f'Renamed to: {mapped_name}'
            result['notes'] = f'Type renamed from {qualified_name} to {mapped_name}'
        
        # Compare base types
        base_old = get_base_type_name(ct_old, NS_26)
        base_new = get_base_type_name(ct_new, NS_3)
        
        if base_old != base_new:
            if base_new:
                result['base_changed'] = f"{base_new} (was: {base_old if base_old else 'none'})"
        
        # Resolve content models
        cm_old_elements, cm_old_attrs = resolve_content_model(
            qualified_name, self.old_types, self.old_attrs, NS_26
        )
        cm_new_elements, cm_new_attrs = resolve_content_model(
            mapped_name if mapped_name else qualified_name, self.new_types, self.new_attrs, NS_3
        )
        
        names_old = set(cm_old_elements.keys())
        names_new = set(cm_new_elements.keys())
        attr_names_old = set(cm_old_attrs.keys())
        attr_names_new = set(cm_new_attrs.keys())
        
        # New elements/attributes
        new_elem_names = names_new - names_old
        for elem_name in sorted(new_elem_names):
            elem = cm_new_elements[elem_name]
            card = self.format_cardinality(elem['minOccurs'], elem['maxOccurs'])
            result['new_elements'].append(f"{elem_name}{card}")
            if elem['minOccurs'] != '0':
                result['compatible'] = 'No'
                result['notes'] += f'; New required element: {elem_name}' if result['notes'] else f'New required element: {elem_name}'
        
        new_attr_names = attr_names_new - attr_names_old
        for attr_name in sorted(new_attr_names):
            attr = cm_new_attrs[attr_name]
            result['new_elements'].append(f"@{attr_name}({attr['use']})")
            if attr['use'] == 'required':
                result['compatible'] = 'No'
                result['notes'] += f'; New required attribute: {attr_name}' if result['notes'] else f'New required attribute: {attr_name}'
        
        # Missing elements/attributes
        missing_elem_names = names_old - names_new
        for elem_name in sorted(missing_elem_names):
            result['missing_elements'].append(elem_name)
            result['compatible'] = 'No'
            result['notes'] += f'; Element removed: {elem_name}' if result['notes'] else f'Element removed: {elem_name}'
        
        missing_attr_names = attr_names_old - attr_names_new
        for attr_name in sorted(missing_attr_names):
            result['missing_elements'].append(f"@{attr_name}")
            result['compatible'] = 'No'
            result['notes'] += f'; Attribute removed: {attr_name}' if result['notes'] else f'Attribute removed: {attr_name}'
        
        # Check common elements
        common_elems = names_old & names_new
        for elem_name in sorted(common_elems):
            e_old = cm_old_elements[elem_name]
            e_new = cm_new_elements[elem_name]
            
            min_old = int(e_old['minOccurs']) if e_old['minOccurs'].isdigit() else 0
            min_new = int(e_new['minOccurs']) if e_new['minOccurs'].isdigit() else 0
            max_old = 999999 if e_old['maxOccurs'] == 'unbounded' else int(e_old['maxOccurs']) if e_old['maxOccurs'].isdigit() else 1
            max_new = 999999 if e_new['maxOccurs'] == 'unbounded' else int(e_new['maxOccurs']) if e_new['maxOccurs'].isdigit() else 1
            
            if min_new < min_old or max_new > max_old:
                card = self.format_cardinality(e_new['minOccurs'], e_new['maxOccurs'])
                result['expanded_cardinality'].append(f"{elem_name}{card}")
            elif min_new > min_old or max_new < max_old:
                card = self.format_cardinality(e_new['minOccurs'], e_new['maxOccurs'])
                result['restricted_cardinality'].append(f"{elem_name}{card}")
                result['compatible'] = 'No'
                result['notes'] += f'; Cardinality restricted: {elem_name}' if result['notes'] else f'Cardinality restricted: {elem_name}'
            
            # Type changes
            if e_old['type'] != e_new['type'] and e_new['type']:
                type_compatible, compat_reason = self.is_type_change_compatible(
                    e_old['type'], e_new['type']
                )
                
                type_note = f"{elem_name}: {e_old['type']} → {e_new['type']}"
                
                if type_compatible:
                    result['type_changes'].append(f"{type_note} [OK]")
                else:
                    result['type_changes'].append(f"{type_note} [INCOMPATIBLE: {compat_reason}]")
                    result['compatible'] = 'No'
                    result['notes'] += f'; Incompatible type change: {elem_name} ({compat_reason})' if result['notes'] else f'Incompatible type change: {elem_name} ({compat_reason})'
        
        # Check attribute use changes
        common_attrs = attr_names_old & attr_names_new
        for attr_name in sorted(common_attrs):
            a_old = cm_old_attrs[attr_name]
            a_new = cm_new_attrs[attr_name]
            
            if a_old['use'] == 'required' and a_new['use'] == 'optional':
                result['expanded_cardinality'].append(f"@{attr_name}(optional)")
            elif a_old['use'] == 'optional' and a_new['use'] == 'required':
                result['restricted_cardinality'].append(f"@{attr_name}(required)")
                result['compatible'] = 'No'
                result['notes'] += f'; Attribute now required: {attr_name}' if result['notes'] else f'Attribute now required: {attr_name}'
        
        # Final note for base type changes
        if result['compatible'] == 'Yes' and result['base_changed']:
            result['notes'] = 'Base type changed but content model compatible (architectural refactoring)'
        
        return result
    
    def compare_simpletype(self, qualified_name: str, mapped_name: Optional[str] = None) -> Dict:
        """Compare a simple type between old and new versions
        
        Args:
            qualified_name: Qualified name like "diggs:SomeSimpleType" or "eml:UnitlessMeasure"
            mapped_name: Optional mapped/renamed qualified name in new version
        """
        result = {
            'name': qualified_name,  # Now includes namespace prefix
            'type_removed': '',
            'base_changed': '',
            'restriction_changed': '',
            'notes': '',
            'compatible': 'Yes'
        }
        
        st_old = self.old_simpletypes.get(qualified_name)
        st_new = self.new_simpletypes.get(mapped_name if mapped_name else qualified_name)
        
        if st_new is None:
            result['type_removed'] = 'Yes'
            result['compatible'] = 'No'
            result['notes'] = 'SimpleType not found in new version'
            return result
        
        result['type_removed'] = 'No'
        
        if mapped_name:
            result['type_removed'] = f'Renamed to: {mapped_name}'
            result['notes'] = f'SimpleType renamed from {qualified_name} to {mapped_name}'
        
        # For simple types, we check if the base type or restrictions changed
        # Use xml.etree.ElementTree (standard library) for serialization
        import xml.etree.ElementTree as ET
        
        old_str = ET.tostring(st_old, encoding='unicode') if st_old is not None else ''
        new_str = ET.tostring(st_new, encoding='unicode') if st_new is not None else ''
        
        if old_str != new_str:
            result['restriction_changed'] = 'Yes'
            result['notes'] += '; Restriction or base type changed' if result['notes'] else 'Restriction or base type changed'
        
        return result
    
    def analyze(self) -> List[Dict]:
        """Run the full compatibility analysis"""
        print("="*80)
        print("DIGGS SCHEMA COMPATIBILITY ANALYZER")
        print("="*80)
        
        # Detect versions
        print("\nDetecting versions...")
        self.old_version = self.detect_version(self.old_dir)
        self.new_version = self.detect_version(self.new_dir)
        print(f"  Old version: {self.old_version}")
        print(f"  New version: {self.new_version}")
        
        # Load type mappings
        self.type_mappings = self.load_mappings()
        
        # Load schemas
        self.old_schemas, self.old_types, self.old_attrs, self.old_simpletypes = self.load_schemas(
            self.old_dir, NS_26
        )
        self.new_schemas, self.new_types, self.new_attrs, self.new_simpletypes = self.load_schemas(
            self.new_dir, NS_3
        )
        
        # Compare types
        print(f"\nComparing {len(self.old_types)} types...")
        results = []
        
        for i, qualified_name in enumerate(sorted(self.old_types.keys()), 1):
            if i % 100 == 0:
                print(f"  Progress: {i}/{len(self.old_types)}")
            
            # qualified_name is already in format "namespace:typename"
            # Try exact match in mappings first
            mapped_name = self.type_mappings.get(qualified_name)
            
            # If not found, try bare name for backward compatibility
            if not mapped_name:
                bare_name = qualified_name.split(':')[-1]
                mapped_name = self.type_mappings.get(bare_name)
                # If found with bare name, qualify it with same namespace as target
                if mapped_name and ':' not in mapped_name:
                    # Extract namespace from old qualified name
                    ns_prefix = qualified_name.split(':')[0]
                    # Check if target exists with same namespace
                    potential_new = f"{ns_prefix}:{mapped_name}"
                    if potential_new in self.new_types:
                        mapped_name = potential_new
            
            result = self.compare_type(qualified_name, mapped_name)
            results.append(result)
        
        # Compare simpleTypes
        print(f"\nComparing {len(self.old_simpletypes)} simpleTypes...")
        simpletype_results = []
        
        for i, qualified_name in enumerate(sorted(self.old_simpletypes.keys()), 1):
            if i % 100 == 0:
                print(f"  Progress: {i}/{len(self.old_simpletypes)}")
            
            # Try exact match in mappings first
            mapped_name = self.type_mappings.get(qualified_name)
            
            # If not found, try bare name for backward compatibility
            if not mapped_name:
                bare_name = qualified_name.split(':')[-1]
                mapped_name = self.type_mappings.get(bare_name)
                if mapped_name and ':' not in mapped_name:
                    ns_prefix = qualified_name.split(':')[0]
                    potential_new = f"{ns_prefix}:{mapped_name}"
                    if potential_new in self.new_simpletypes:
                        mapped_name = potential_new
            
            result = self.compare_simpletype(qualified_name, mapped_name)
            simpletype_results.append(result)
        
        # Clear cache to free memory
        self.type_compat_cache.clear()
        
        # Statistics
        removed = sum(1 for r in results if r['type_removed'] == 'Yes')
        renamed = sum(1 for r in results if r['type_removed'].startswith('Renamed'))
        compatible = sum(1 for r in results if r['compatible'] == 'Yes')
        incompatible = sum(1 for r in results if r['compatible'] == 'No' and r['type_removed'] != 'Yes')
        
        print(f"\nResults:")
        print(f"  Compatible: {compatible} ({100*compatible/len(results):.1f}%)")
        print(f"  Incompatible: {incompatible} ({100*incompatible/len(results):.1f}%)")
        print(f"  Renamed: {renamed}")
        print(f"  Removed: {removed} ({100*removed/len(results):.1f}%)")
        
        return results, simpletype_results
    
    def generate_excel_report(self, results: List[Dict], simpletype_results: List[Dict], output_file: str):
        """Generate Excel workbook with analysis results"""
        print(f"\nGenerating Excel report: {output_file}")
        
        wb = Workbook()
        
        # Sheet 1: ComplexTypes comparison
        ws = wb.active
        ws.title = "ComplexTypes"
        
        headers = [
            'Complex Type Name', 'Type Removed',
            'BaseType Changed To', 'Element Order Changed', 'New Element/Attr',
            'Missing Element/Attr', 'Element/Attr Cardinality Expanded',
            'Element/Attr Cardinality Restricted', 'Element/Attr Type Changed',
            'Notes', 'Backward Compatible'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        for row_idx, r in enumerate(results, 2):
            ws.cell(row_idx, 1).value = r['name']  # Now includes namespace prefix
            ws.cell(row_idx, 2).value = r['type_removed']
            ws.cell(row_idx, 3).value = r['base_changed']
            ws.cell(row_idx, 4).value = r['structure_changes']
            ws.cell(row_idx, 5).value = '\n'.join(r['new_elements'][:10])
            ws.cell(row_idx, 6).value = '\n'.join(r['missing_elements'][:10])
            ws.cell(row_idx, 7).value = '\n'.join(r['expanded_cardinality'][:10])
            ws.cell(row_idx, 8).value = '\n'.join(r['restricted_cardinality'][:10])
            ws.cell(row_idx, 9).value = '\n'.join(r['type_changes'][:10])
            ws.cell(row_idx, 10).value = r['notes']
            ws.cell(row_idx, 11).value = r['compatible']
            
            compat_cell = ws.cell(row_idx, 11)
            if r['compatible'] == 'No':
                compat_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            else:
                compat_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            
            for col in range(1, 12):
                ws.cell(row_idx, col).alignment = Alignment(wrap_text=True, vertical='top')
        
        ws.column_dimensions['A'].width = 50  # Wider for qualified names
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 30
        ws.column_dimensions['H'].width = 30
        ws.column_dimensions['I'].width = 40
        ws.column_dimensions['J'].width = 50
        ws.column_dimensions['K'].width = 18
        ws.freeze_panes = 'A2'
        
        # Sheet 2: Type Lists
        ws2 = wb.create_sheet('Type Lists')
        
        ws2.cell(1, 1).value = f"{self.old_version} Types NOT in {self.new_version}"
        ws2.cell(1, 1).font = Font(bold=True, size=14)
        ws2.cell(1, 2).value = f"{self.new_version} Types NOT in {self.old_version}"
        ws2.cell(1, 2).font = Font(bold=True, size=14)
        
        old_names = set(self.old_types.keys())  # Already qualified
        new_names = set(self.new_types.keys())  # Already qualified
        
        # Build sets of renamed types (to exclude from lists)
        # Old types that were mapped to new types (renamed, not truly removed)
        mapped_old_types = set()
        for old_type, new_type in self.type_mappings.items():
            # old_type might be qualified (eml:X) or bare (X)
            # Check if it exists in old_names (which are qualified)
            if old_type in old_names:
                mapped_old_types.add(old_type)
            else:
                # Try to find it by bare name
                bare_old = old_type.split(':')[-1]
                for qualified_old in old_names:
                    if qualified_old.endswith(':' + bare_old) or qualified_old == bare_old:
                        mapped_old_types.add(qualified_old)
                        break
        
        # New types that are targets of mappings (renamed from old, not truly new)
        mapped_to_new_types = set()
        for old_type, new_type in self.type_mappings.items():
            if new_type in new_names:
                mapped_to_new_types.add(new_type)
            else:
                # Try to find it by bare name
                bare_new = new_type.split(':')[-1]
                for qualified_new in new_names:
                    if qualified_new.endswith(':' + bare_new) or qualified_new == bare_new:
                        mapped_to_new_types.add(qualified_new)
                        break
        
        # Filter out renamed types - only show truly removed/new types
        old_not_in_new = sorted((old_names - new_names) - mapped_old_types)
        new_not_in_old = sorted((new_names - old_names) - mapped_to_new_types)
        
        ws2.cell(2, 1).value = f"Count: {len(old_not_in_new)} (excludes renamed types)"
        ws2.cell(2, 1).font = Font(bold=True)
        for idx, qualified_name in enumerate(old_not_in_new, 3):
            ws2.cell(idx, 1).value = qualified_name
        
        ws2.cell(2, 2).value = f"Count: {len(new_not_in_old)} (excludes renamed types)"
        ws2.cell(2, 2).font = Font(bold=True)
        for idx, qualified_name in enumerate(new_not_in_old, 3):
            ws2.cell(idx, 2).value = qualified_name
        
        ws2.column_dimensions['A'].width = 50
        ws2.column_dimensions['B'].width = 50
        
        # Add simpleTypes section
        ws2.cell(1, 4).value = f"{self.old_version} SimpleTypes NOT in {self.new_version}"
        ws2.cell(1, 4).font = Font(bold=True, size=14)
        ws2.cell(1, 5).value = f"{self.new_version} SimpleTypes NOT in {self.old_version}"
        ws2.cell(1, 5).font = Font(bold=True, size=14)
        
        old_st_names = set(self.old_simpletypes.keys())  # Already qualified
        new_st_names = set(self.new_simpletypes.keys())  # Already qualified
        
        # Build sets of renamed simpleTypes (to exclude from lists)
        mapped_old_simpletypes = set()
        for old_type, new_type in self.type_mappings.items():
            if old_type in old_st_names:
                mapped_old_simpletypes.add(old_type)
            else:
                bare_old = old_type.split(':')[-1]
                for qualified_old in old_st_names:
                    if qualified_old.endswith(':' + bare_old) or qualified_old == bare_old:
                        mapped_old_simpletypes.add(qualified_old)
                        break
        
        mapped_to_new_simpletypes = set()
        for old_type, new_type in self.type_mappings.items():
            if new_type in new_st_names:
                mapped_to_new_simpletypes.add(new_type)
            else:
                bare_new = new_type.split(':')[-1]
                for qualified_new in new_st_names:
                    if qualified_new.endswith(':' + bare_new) or qualified_new == bare_new:
                        mapped_to_new_simpletypes.add(qualified_new)
                        break
        
        # Filter out renamed simpleTypes
        old_st_not_in_new = sorted((old_st_names - new_st_names) - mapped_old_simpletypes)
        new_st_not_in_old = sorted((new_st_names - old_st_names) - mapped_to_new_simpletypes)
        
        ws2.cell(2, 4).value = f"Count: {len(old_st_not_in_new)} (excludes renamed types)"
        ws2.cell(2, 4).font = Font(bold=True)
        for idx, qualified_name in enumerate(old_st_not_in_new, 3):
            ws2.cell(idx, 4).value = qualified_name
        
        ws2.cell(2, 5).value = f"Count: {len(new_st_not_in_old)} (excludes renamed types)"
        ws2.cell(2, 5).font = Font(bold=True)
        for idx, qualified_name in enumerate(new_st_not_in_old, 3):
            ws2.cell(idx, 5).value = qualified_name
        
        ws2.column_dimensions['D'].width = 50
        ws2.column_dimensions['E'].width = 50
        
        # Sheet 3: Type Mappings Applied
        ws3 = wb.create_sheet('Type Mappings')
        ws3.cell(1, 1).value = f"{self.old_version} Type Name"
        ws3.cell(1, 1).font = Font(bold=True)
        ws3.cell(1, 2).value = f"{self.new_version} Type Name"
        ws3.cell(1, 2).font = Font(bold=True)
        
        row = 2
        for old_name, new_name in sorted(self.type_mappings.items()):
            ws3.cell(row, 1).value = old_name
            ws3.cell(row, 2).value = new_name
            row += 1
        
        ws3.column_dimensions['A'].width = 50
        ws3.column_dimensions['B'].width = 50
        
        # Sheet 4: SimpleTypes comparison
        ws4 = wb.create_sheet('SimpleTypes')
        
        st_headers = [
            'Simple Type Name', 'Type Removed',
            'Base/Restriction Changed', 'Notes', 'Backward Compatible'
        ]
        
        for col, header in enumerate(st_headers, 1):
            cell = ws4.cell(1, col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        for row_idx, r in enumerate(simpletype_results, 2):
            ws4.cell(row_idx, 1).value = r['name']  # Now includes namespace prefix
            ws4.cell(row_idx, 2).value = r['type_removed']
            ws4.cell(row_idx, 3).value = r['restriction_changed']
            ws4.cell(row_idx, 4).value = r['notes']
            ws4.cell(row_idx, 5).value = r['compatible']
            
            compat_cell = ws4.cell(row_idx, 5)
            if r['compatible'] == 'No':
                compat_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            else:
                compat_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            
            for col in range(1, 6):
                ws4.cell(row_idx, col).alignment = Alignment(wrap_text=True, vertical='top')
        
        ws4.column_dimensions['A'].width = 50
        ws4.column_dimensions['B'].width = 25
        ws4.column_dimensions['C'].width = 25
        ws4.column_dimensions['D'].width = 50
        ws4.column_dimensions['E'].width = 18
        ws4.freeze_panes = 'A2'
        
        wb.save(output_file)
        print(f"  ✓ Report saved: {output_file}")


def main():
    """Main entry point"""
    parser = argparse.ArgumentParser(
        description='DIGGS Schema Compatibility Analyzer',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Basic usage with directories
  python diggs_compatibility_analyzer.py \\
    --old-dir /path/to/diggs_v2.6 \\
    --new-dir /path/to/diggs_v3.0 \\
    --output compatibility_report.xlsx

  # With type mappings file
  python diggs_compatibility_analyzer.py \\
    --old-dir /path/to/diggs_v2.6 \\
    --new-dir /path/to/diggs_v3.0 \\
    --mappings type_mappings.txt \\
    --output compatibility_report.xlsx

Type mappings file format (tab-separated):
  OldTypeName<TAB>NewTypeName
  AbstractProgramPropertyType<TAB>ProgramPropertyType
  BearingType<TAB>BearingMeasureType
        """
    )
    
    parser.add_argument(
        '--old-dir',
        required=True,
        help='Directory containing old version DIGGS schema files'
    )
    
    parser.add_argument(
        '--new-dir',
        required=True,
        help='Directory containing new version DIGGS schema files'
    )
    
    parser.add_argument(
        '--mappings',
        help='Path to type mappings file (tab-separated: old_type<TAB>new_type)'
    )
    
    parser.add_argument(
        '--output',
        default='diggs_compatibility_analysis.xlsx',
        help='Output Excel file name (default: diggs_compatibility_analysis.xlsx)'
    )
    
    args = parser.parse_args()
    
    try:
        # Create analyzer
        analyzer = DIGGSCompatibilityAnalyzer(
            args.old_dir,
            args.new_dir,
            args.mappings
        )
        
        # Run analysis
        results, simpletype_results = analyzer.analyze()
        
        # Generate report
        analyzer.generate_excel_report(results, simpletype_results, args.output)
        
        print("\n" + "="*80)
        print("✓ Analysis complete!")
        print("="*80)
        
    except Exception as e:
        print(f"\nError: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
