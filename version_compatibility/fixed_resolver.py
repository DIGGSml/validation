#!/usr/bin/env python3
"""
Fixed content model resolver - properly handles nested sequences
"""

import xml.etree.ElementTree as ET
from collections import OrderedDict
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

NS_26 = {
    'xs': 'http://www.w3.org/2001/XMLSchema',
    'diggs': 'http://diggsml.org/schemas/2.6',
    'gml': 'http://www.opengis.net/gml/3.2',
    'eml': 'http://www.energistics.org/energyml/data/commonv2'
}

NS_3 = {
    'xs': 'http://www.w3.org/2001/XMLSchema',
    'diggs': 'http://diggsml.org/schemas/3.0',
    'gml': 'http://www.opengis.net/gml/3.2',
    'eml': 'http://www.energistics.org/energyml/data/commonv2'
}

def parse_schema(filepath, namespaces):
    try:
        tree = ET.parse(filepath)
        return tree.getroot()
    except Exception as e:
        print(f"Warning: Could not parse {filepath}: {e}")
        return None

def get_all_types(schemas, namespaces):
    all_types = {}
    for schema in schemas:
        if schema is not None:
            for ct in schema.findall('.//xs:complexType', namespaces):
                name = ct.get('name')
                if name:
                    all_types[name] = ct
    return all_types

def get_all_attributes(schemas, namespaces):
    all_attrs = {}
    for schema in schemas:
        if schema is not None:
            for attr in schema.findall('.//xs:attribute', namespaces):
                name = attr.get('name')
                if name:
                    all_attrs[name] = attr
    return all_attrs

def get_all_simpletypes(schemas, namespaces):
    """Collect all simpleType definitions from schemas"""
    all_simpletypes = {}
    for schema in schemas:
        if schema is not None:
            for st in schema.findall('.//xs:simpleType', namespaces):
                name = st.get('name')
                if name:
                    all_simpletypes[name] = st
    return all_simpletypes

def clean_type_name(name):
    if not name:
        return ''
    name = name.replace('diggs:', '')
    return name

def get_base_type_name(complex_type, namespaces):
    extension = complex_type.find('.//xs:extension', namespaces)
    if extension is not None:
        base = extension.get('base')
        if base:
            return clean_type_name(base)
    
    restriction = complex_type.find('.//xs:restriction', namespaces)
    if restriction is not None:
        base = restriction.get('base')
        if base:
            return clean_type_name(base)
    
    return None

def extract_elements_from_node(node, namespaces):
    """
    Recursively extract all elements from a node, handling nested sequences and choices.
    This is the key fix - it traverses nested structures properly.
    """
    elements = []
    
    # Get all sequences at any level under this node
    for seq in node.findall('.//xs:sequence', namespaces):
        # Get direct element children of this sequence
        for elem in seq.findall('xs:element', namespaces):
            elem_info = {
                'name': elem.get('name') or clean_type_name(elem.get('ref', '')),
                'type': clean_type_name(elem.get('type', '')),
                'minOccurs': elem.get('minOccurs', '1'),
                'maxOccurs': elem.get('maxOccurs', '1'),
            }
            # Avoid duplicates
            if not any(e['name'] == elem_info['name'] for e in elements):
                elements.append(elem_info)
    
    # Get all choices at any level
    for choice in node.findall('.//xs:choice', namespaces):
        for elem in choice.findall('xs:element', namespaces):
            elem_info = {
                'name': elem.get('name') or clean_type_name(elem.get('ref', '')),
                'type': clean_type_name(elem.get('type', '')),
                'minOccurs': elem.get('minOccurs', '1'),
                'maxOccurs': elem.get('maxOccurs', '1'),
            }
            if not any(e['name'] == elem_info['name'] for e in elements):
                elements.append(elem_info)
    
    return elements

def extract_local_elements(complex_type, namespaces):
    """Extract elements defined locally in this type"""
    # Check for extension or restriction
    extension = complex_type.find('.//xs:extension', namespaces)
    is_extension = extension is not None
    
    restriction = complex_type.find('.//xs:restriction', namespaces)
    
    # Extract from the appropriate content model
    if extension is not None:
        elements = extract_elements_from_node(extension, namespaces)
    elif restriction is not None:
        elements = extract_elements_from_node(restriction, namespaces)
    else:
        elements = extract_elements_from_node(complex_type, namespaces)
    
    return elements, is_extension

def resolve_attribute_ref(attr_ref, all_attrs, namespaces):
    attr_name = attr_ref.replace('gml:', '').replace('diggs:', '').replace('xml:', '')
    
    if attr_name in all_attrs:
        attr_def = all_attrs[attr_name]
        return {
            'name': attr_ref,
            'type': clean_type_name(attr_def.get('type', '')),
            'use': 'optional',
        }
    
    return None

def extract_local_attributes(complex_type, all_attrs, namespaces):
    attributes = []
    
    for attr in complex_type.findall('.//xs:attribute', namespaces):
        ref = attr.get('ref')
        if ref:
            resolved_attr = resolve_attribute_ref(clean_type_name(ref), all_attrs, namespaces)
            if resolved_attr:
                use_override = attr.get('use')
                if use_override:
                    resolved_attr['use'] = use_override
                attributes.append(resolved_attr)
        else:
            attr_info = {
                'name': attr.get('name'),
                'type': clean_type_name(attr.get('type', '')),
                'use': attr.get('use', 'optional'),
            }
            attributes.append(attr_info)
    
    return attributes

def resolve_content_model(type_name, all_types, all_attrs, namespaces, visited=None, depth=0):
    """Recursively resolve the complete content model"""
    if visited is None:
        visited = set()
    
    if depth > 20:
        return {}, {}
    
    if type_name in visited:
        return {}, {}
    
    visited.add(type_name)
    
    if type_name.startswith('xs:') or not type_name:
        return {}, {}
    
    lookup_name = type_name.replace('gml:', '').replace('diggs:', '')
    
    complex_type = all_types.get(lookup_name)
    if complex_type is None:
        return {}, {}
    
    base_type_name = get_base_type_name(complex_type, namespaces)
    local_elements, is_extension = extract_local_elements(complex_type, namespaces)
    local_attributes = extract_local_attributes(complex_type, all_attrs, namespaces)
    
    elements_dict = OrderedDict()
    attributes_dict = OrderedDict()
    
    # If extension, inherit base content
    if is_extension and base_type_name:
        base_elements, base_attributes = resolve_content_model(
            base_type_name, all_types, all_attrs, namespaces, visited.copy(), depth + 1
        )
        elements_dict.update(base_elements)
        attributes_dict.update(base_attributes)
    
    # If restriction, attributes pass through
    elif base_type_name and not is_extension:
        _, base_attributes = resolve_content_model(
            base_type_name, all_types, all_attrs, namespaces, visited.copy(), depth + 1
        )
        attributes_dict.update(base_attributes)
    
    # Add/override local elements
    for elem in local_elements:
        elements_dict[elem['name']] = elem
    
    # Add/override local attributes
    for attr in local_attributes:
        attributes_dict[attr['name']] = attr
    
    return elements_dict, attributes_dict

def content_models_match(cm1_elements, cm1_attrs, cm2_elements, cm2_attrs, strict=True):
    differences = []
    
    elem_names_1 = set(cm1_elements.keys())
    elem_names_2 = set(cm2_elements.keys())
    
    removed_elems = elem_names_1 - elem_names_2
    added_elems = elem_names_2 - elem_names_1
    
    if removed_elems:
        differences.append(f"Elements removed: {', '.join(sorted(list(removed_elems))[:5])}")
    if strict and added_elems:
        differences.append(f"Elements added: {', '.join(sorted(list(added_elems))[:5])}")
    
    for elem_name in elem_names_1 & elem_names_2:
        e1 = cm1_elements[elem_name]
        e2 = cm2_elements[elem_name]
        
        min1 = int(e1['minOccurs']) if e1['minOccurs'].isdigit() else 0
        min2 = int(e2['minOccurs']) if e2['minOccurs'].isdigit() else 0
        
        if min2 > min1:
            differences.append(f"{elem_name}: minOccurs increased")
        
        max1 = 999999 if e1['maxOccurs'] == 'unbounded' else int(e1['maxOccurs']) if e1['maxOccurs'].isdigit() else 1
        max2 = 999999 if e2['maxOccurs'] == 'unbounded' else int(e2['maxOccurs']) if e2['maxOccurs'].isdigit() else 1
        
        if max2 < max1:
            differences.append(f"{elem_name}: maxOccurs decreased")
    
    attr_names_1 = set(cm1_attrs.keys())
    attr_names_2 = set(cm2_attrs.keys())
    
    removed_attrs = attr_names_1 - attr_names_2
    added_attrs = attr_names_2 - attr_names_1
    
    if removed_attrs:
        differences.append(f"Attributes removed: {', '.join(sorted(list(removed_attrs))[:5])}")
    if strict and added_attrs:
        differences.append(f"Attributes added: {', '.join(sorted(list(added_attrs))[:5])}")
    
    for attr_name in attr_names_1 & attr_names_2:
        a1 = cm1_attrs[attr_name]
        a2 = cm2_attrs[attr_name]
        
        if a1['use'] == 'optional' and a2['use'] == 'required':
            differences.append(f"@{attr_name}: changed to required")
    
    return len(differences) == 0, differences

def main():
    print("=" * 80)
    print("Fixed Content Model Resolver - Handles Nested Sequences")
    print("=" * 80)
    
    print("\nLoading existing comparison workbook...")
    wb = load_workbook('/mnt/user-data/outputs/DIGGS_Detailed_Comparison_v26_to_v30_MAPPED.xlsx')
    ws = wb['ComplexTypes']
    
    print("\nLoading schemas...")
    
    v26_schemas = [
        parse_schema('/mnt/user-data/uploads/Kernel.xsd', NS_26),
        parse_schema('/mnt/user-data/uploads/gml3_2Profile_diggs.xsd', NS_26),
    ]
    all_types_26 = get_all_types(v26_schemas, NS_26)
    all_attrs_26 = get_all_attributes(v26_schemas, NS_26)
    print(f"    Found {len(all_types_26)} types and {len(all_attrs_26)} global attributes in v2.6")
    
    v3_files = [
        'AbstractTypes.xsd', 'Common.xsd', 'Core.xsd',
        'LinearReferencing.xsd', 'LocalReferencing.xsd', 'Observation.xsd',
        'Groups.xsd', 'ExtSamplingFeatures.xsd', 'DeletedElements.xsd',
        'Geophysics.xsd', 'Diggs.xsd', 'DiggsCore.xsd',
        'Grouting.xsd', 'ExtProcedures.xsd', 'DeepFoundation.xsd',
        'gml3_2Profile_diggs.xsd'
    ]
    
    v3_schemas = []
    for filename in v3_files:
        schema = parse_schema(f'/mnt/user-data/uploads/{filename}', NS_3)
        if schema is not None:
            v3_schemas.append(schema)
    
    all_types_3 = get_all_types(v3_schemas, NS_3)
    all_attrs_3 = get_all_attributes(v3_schemas, NS_3)
    print(f"    Found {len(all_types_3)} types and {len(all_attrs_3)} global attributes in v3.0")
    
    print("\nTesting fix on AbstractComponentObjectType...")
    cm26_elements, cm26_attrs = resolve_content_model(
        'AbstractComponentObjectType', all_types_26, all_attrs_26, NS_26
    )
    cm3_elements, cm3_attrs = resolve_content_model(
        'AbstractComponentObjectType', all_types_3, all_attrs_3, NS_3
    )
    
    print(f"  v2.6 content: {len(cm26_elements)} elements, {len(cm26_attrs)} attributes")
    print(f"  v2.6 elements: {', '.join(cm26_elements.keys())}")
    print(f"  v3.0 content: {len(cm3_elements)} elements, {len(cm3_attrs)} attributes")
    print(f"  v3.0 elements: {', '.join(cm3_elements.keys())}")
    
    match, diffs = content_models_match(cm26_elements, cm26_attrs, cm3_elements, cm3_attrs, strict=False)
    if match:
        print("  ✓ Content models MATCH!")
    else:
        print("  ✗ Content models DIFFER:")
        for diff in diffs:
            print(f"    - {diff}")
    
    print("\nFinding all types marked as incompatible...")
    types_to_check = []
    
    for row in range(2, ws.max_row + 1):
        type_name = ws.cell(row, 1).value
        compatible = ws.cell(row, 11).value
        
        if compatible == 'No':
            types_to_check.append(row)
    
    print(f"  Found {len(types_to_check)} types marked as incompatible")
    
    updated_count = 0
    checked_count = 0
    
    for row in types_to_check:
        type_name = ws.cell(row, 1).value
        type_removed = ws.cell(row, 2).value
        if type_removed == 'Yes':
            continue
        
        checked_count += 1
        
        if checked_count <= 5 or checked_count % 10 == 0:
            print(f"\n  [{checked_count}/{len(types_to_check)}] Analyzing: {type_name}")
        
        cm26_elements, cm26_attrs = resolve_content_model(
            type_name, all_types_26, all_attrs_26, NS_26
        )
        cm3_elements, cm3_attrs = resolve_content_model(
            type_name, all_types_3, all_attrs_3, NS_3
        )
        
        if checked_count <= 5:
            print(f"    v2.6 content: {len(cm26_elements)} elements, {len(cm26_attrs)} attributes")
            print(f"    v3.0 content: {len(cm3_elements)} elements, {len(cm3_attrs)} attributes")
        
        match, differences = content_models_match(
            cm26_elements, cm26_attrs, cm3_elements, cm3_attrs, strict=False
        )
        
        if match:
            if checked_count <= 5:
                print(f"    ✓ Content models COMPATIBLE - updating")
            ws.cell(row, 11).value = 'Yes'
            ws.cell(row, 11).fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            
            current_notes = ws.cell(row, 10).value or ''
            if current_notes:
                ws.cell(row, 10).value = current_notes + '; Content model analysis confirms compatibility'
            else:
                ws.cell(row, 10).value = 'Content model analysis confirms compatibility'
            
            updated_count += 1
        else:
            if checked_count <= 5:
                print(f"    ✗ Content models INCOMPATIBLE:")
                for diff in differences[:2]:
                    print(f"      - {diff}")
    
    output_file = '/mnt/user-data/outputs/DIGGS_v26_to_v30_Comparison_FINAL.xlsx'
    wb.save(output_file)
    
    print("\n" + "=" * 80)
    print(f"Analysis complete!")
    print(f"  Types checked: {checked_count}")
    print(f"  Compatibility updated: {updated_count}")
    print(f"  Output: {output_file}")
    print("=" * 80)

if __name__ == '__main__':
    main()
